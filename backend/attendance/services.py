import calendar
import hashlib
import re
from datetime import date
from decimal import Decimal

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    AttendancePolicy,
    AttendanceResult,
    CrossDaySuspicion,
    Employee,
    ImportBatch,
    RawPunchDay,
)


TIME_TOKEN = re.compile(r"^(?P<next>次日)?(?P<hour>\d{1,2}):(?P<minute>\d{2})$")


def normalize_name(value):
    return re.sub(r"\s+|[（(]已离职[）)]", "", str(value or "")).strip()


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def file_sha256(uploaded_file):
    digest = hashlib.sha256()
    position = uploaded_file.tell()
    uploaded_file.seek(0)
    for chunk in iter(lambda: uploaded_file.read(1024 * 1024), b""):
        digest.update(chunk)
    uploaded_file.seek(position)
    return digest.hexdigest()


def parse_punches(value):
    raw = cell_text(value)
    if not raw or raw == "-":
        return []
    punches = []
    for token in re.split(r"[\r\n]+", raw):
        token = token.strip()
        if not token or token == "-":
            continue
        match = TIME_TOKEN.match(token)
        if not match:
            punches.append({"text": token, "minutes": None, "next_day": False, "valid": False})
            continue
        minutes = int(match.group("hour")) * 60 + int(match.group("minute"))
        next_day = bool(match.group("next"))
        punches.append(
            {
                "text": token,
                "minutes": minutes + (1440 if next_day else 0),
                "clock_minutes": minutes,
                "next_day": next_day,
                "valid": True,
            }
        )
    return punches


def _employee_maps():
    employees = list(Employee.objects.select_related("attendance_policy").prefetch_related("tags"))
    by_no = {employee.employee_no.strip(): employee for employee in employees if employee.employee_no.strip()}
    by_name = {}
    by_alias = {}
    for employee in employees:
        by_name.setdefault(normalize_name(employee.name), employee)
        for alias in employee.aliases or []:
            by_alias.setdefault(normalize_name(alias), employee)
    return employees, by_no, by_name, by_alias


def _match_employee(employee_no, source_name, by_no, by_name, by_alias):
    if employee_no and employee_no in by_no:
        return by_no[employee_no], RawPunchDay.MatchStatus.EMPLOYEE_NO
    normalized = normalize_name(source_name)
    if normalized in by_name:
        return by_name[normalized], RawPunchDay.MatchStatus.NAME
    if normalized in by_alias:
        return by_alias[normalized], RawPunchDay.MatchStatus.ALIAS
    return None, RawPunchDay.MatchStatus.UNMATCHED


def _locate_punch_sheet(workbook):
    if "飞书打卡" in workbook.sheetnames:
        return workbook["飞书打卡"]
    for sheet in workbook.worksheets:
        headers = [cell_text(cell.value) for cell in sheet[1][:4]]
        if headers[:2] == ["姓名", "组织名称"] and "工号" in headers:
            return sheet
    raise ValueError("没有找到打卡明细页：需要包含“姓名、组织名称、工号、考勤规则”表头")


def _day_columns(sheet, year, month):
    max_day = calendar.monthrange(year, month)[1]
    columns = {}
    for cell in sheet[1]:
        header = cell_text(cell.value)
        match = re.match(r"^(\d{1,2})", header)
        if not match:
            continue
        day = int(match.group(1))
        if 1 <= day <= max_day:
            columns[day] = cell.column
    if not columns:
        raise ValueError("打卡明细页没有识别到日期列")
    return columns


@transaction.atomic
def process_import_batch(batch):
    batch.status = ImportBatch.Status.PROCESSING
    batch.error_message = ""
    batch.save(update_fields=["status", "error_message"])
    try:
        workbook = load_workbook(batch.source_file.path, data_only=True, read_only=False)
        sheet = _locate_punch_sheet(workbook)
        day_columns = _day_columns(sheet, batch.year, batch.month)
        employees, by_no, by_name, by_alias = _employee_maps()
        del employees

        batch.raw_days.all().delete()
        source_rows = []
        total_rows = 0
        matched_rows = 0
        unmatched_rows = 0

        for row_number in range(2, sheet.max_row + 1):
            source_name = cell_text(sheet.cell(row_number, 1).value)
            if not source_name:
                continue
            total_rows += 1
            employee_no = cell_text(sheet.cell(row_number, 3).value)
            employee, match_status = _match_employee(
                employee_no,
                source_name,
                by_no,
                by_name,
                by_alias,
            )
            if employee:
                matched_rows += 1
            else:
                unmatched_rows += 1
            organization = cell_text(sheet.cell(row_number, 2).value)
            attendance_rule = cell_text(sheet.cell(row_number, 4).value)
            days_for_row = {}
            for day, column in day_columns.items():
                work_date = date(batch.year, batch.month, day)
                raw_value = cell_text(sheet.cell(row_number, column).value)
                punches = parse_punches(raw_value)
                raw_day = RawPunchDay.objects.create(
                    batch=batch,
                    employee=employee,
                    source_row=row_number,
                    employee_no=employee_no,
                    source_name=source_name,
                    organization=organization,
                    attendance_rule=attendance_rule,
                    work_date=work_date,
                    raw_value=raw_value,
                    punches=punches,
                    has_punch=bool(punches),
                    effective_has_punch=bool(punches),
                    match_status=match_status,
                )
                days_for_row[day] = raw_day
            source_rows.append(days_for_row)

        suspicion_count = 0
        for days_for_row in source_rows:
            for day in sorted(days_for_row):
                if day <= 1:
                    continue
                current = days_for_row[day]
                previous = days_for_row.get(day - 1)
                if not previous or len(current.punches) != 1 or not previous.punches:
                    continue
                punch = current.punches[0]
                if not punch.get("valid") or punch.get("next_day"):
                    continue
                policy_cutoff = (
                    current.employee.attendance_policy.cross_day_cutoff_minutes
                    if current.employee and current.employee.attendance_policy
                    else 180
                )
                if punch.get("clock_minutes", 10_000) > policy_cutoff:
                    continue
                previous_has_late_punch = any(
                    item.get("valid")
                    and not item.get("next_day")
                    and item.get("clock_minutes", 0) >= 12 * 60
                    for item in previous.punches
                )
                if not previous_has_late_punch:
                    continue
                current.is_cross_day_suspicion = True
                current.effective_has_punch = False
                current.save(update_fields=["is_cross_day_suspicion", "effective_has_punch"])
                CrossDaySuspicion.objects.create(
                    batch=batch,
                    raw_day=current,
                    employee=current.employee,
                    previous_date=previous.work_date,
                    work_date=current.work_date,
                    punch_text=punch["text"],
                    reason=f"当天仅一条凌晨打卡，且前一天存在下午或晚间打卡（截止 {policy_cutoff // 60:02d}:{policy_cutoff % 60:02d}）",
                )
                suspicion_count += 1

        recalculate_batch(batch)
        batch.status = ImportBatch.Status.COMPLETED
        batch.total_rows = total_rows
        batch.matched_rows = matched_rows
        batch.unmatched_rows = unmatched_rows
        batch.suspicion_count = suspicion_count
        batch.completed_at = timezone.now()
        batch.save(
            update_fields=[
                "status",
                "total_rows",
                "matched_rows",
                "unmatched_rows",
                "suspicion_count",
                "completed_at",
            ]
        )
    except Exception as exc:
        batch.status = ImportBatch.Status.FAILED
        batch.error_message = str(exc)
        batch.save(update_fields=["status", "error_message"])
        raise
    return batch


def _decimal(value):
    return Decimal(str(value or 0))


@transaction.atomic
def recalculate_result(result):
    batch = result.batch
    employee = result.employee
    policy = employee.attendance_policy
    mode = policy.mode if policy else AttendancePolicy.Mode.STANDARD
    punch_days = RawPunchDay.objects.filter(
        batch=batch,
        employee=employee,
        effective_has_punch=True,
    ).count()
    due_days = employee.expected_days_override or batch.default_expected_days
    days_in_month = calendar.monthrange(batch.year, batch.month)[1]
    rest_days = max(Decimal("0"), _decimal(days_in_month) - _decimal(due_days))
    if mode in {AttendancePolicy.Mode.EXEMPT, AttendancePolicy.Mode.FLEXIBLE}:
        base_actual = _decimal(due_days)
        base_rule = "免考勤/弹性工作：按应出勤天数正常计薪"
    else:
        base_actual = _decimal(punch_days)
        base_rule = "标准规则：空白或“-”为休息，有打卡为出勤；疑似跨日未确认前不计入当天"
    actual_days = base_actual + _decimal(result.adjustment_days)
    pending_suspicions = CrossDaySuspicion.objects.filter(
        batch=batch,
        employee=employee,
        status=CrossDaySuspicion.Status.PENDING,
    ).count()
    result.punch_days = punch_days
    result.due_days = due_days
    result.rest_days = rest_days
    result.actual_days = actual_days
    if result.status != AttendanceResult.Status.APPROVED:
        result.status = (
            AttendanceResult.Status.REVIEW
            if pending_suspicions or (mode == AttendancePolicy.Mode.STANDARD and actual_days != _decimal(due_days))
            else AttendanceResult.Status.NORMAL
        )
    result.rule_trace = {
        "version": "v1",
        "policy_mode": mode,
        "base_rule": base_rule,
        "effective_punch_days": punch_days,
        "due_days": float(due_days),
        "adjustment_days": float(result.adjustment_days),
        "pending_cross_day_suspicions": pending_suspicions,
        "actual_formula": "base_actual + adjustment_days",
    }
    result.save()
    return result


@transaction.atomic
def recalculate_batch(batch):
    month_start = date(batch.year, batch.month, 1)
    active_employees = Employee.objects.filter(active=True).filter(
        models_join_filter(month_start)
    ).select_related("attendance_policy")
    for employee in active_employees:
        result, _ = AttendanceResult.objects.get_or_create(batch=batch, employee=employee)
        recalculate_result(result)


def models_join_filter(month_start):
    from django.db.models import Q

    return Q(join_date__isnull=True) | Q(join_date__lte=month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1]))


@transaction.atomic
def resolve_cross_day(suspicion, resolution, user):
    if resolution not in {
        CrossDaySuspicion.Status.ASSIGN_PREVIOUS,
        CrossDaySuspicion.Status.KEEP_CURRENT,
    }:
        raise ValueError("无效的审核结果")
    suspicion.status = resolution
    suspicion.reviewed_by = user
    suspicion.reviewed_at = timezone.now()
    suspicion.save(update_fields=["status", "reviewed_by", "reviewed_at"])
    suspicion.raw_day.effective_has_punch = resolution == CrossDaySuspicion.Status.KEEP_CURRENT
    suspicion.raw_day.save(update_fields=["effective_has_punch"])
    if suspicion.employee:
        result = AttendanceResult.objects.get(batch=suspicion.batch, employee=suspicion.employee)
        recalculate_result(result)
    return suspicion

