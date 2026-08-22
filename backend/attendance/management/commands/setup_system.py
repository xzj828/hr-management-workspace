from datetime import date, datetime
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from attendance.models import AccountProfile, AttendancePolicy, Employee, EmployeeTag
from attendance.services import cell_text, normalize_name


POLICIES = [
    {
        "code": "standard",
        "name": "标准考勤",
        "mode": AttendancePolicy.Mode.STANDARD,
        "description": "空白或“-”为休息，有打卡为出勤；跨日疑似先审核。",
    },
    {
        "code": "flexible",
        "name": "弹性工作",
        "mode": AttendancePolicy.Mode.FLEXIBLE,
        "description": "保留打卡记录和异常提醒，按应出勤天数正常计薪。",
    },
    {
        "code": "exempt",
        "name": "免考勤",
        "mode": AttendancePolicy.Mode.EXEMPT,
        "description": "不以打卡作为工资依据，按应出勤天数正常计薪。",
    },
    {
        "code": "part_time",
        "name": "兼职",
        "mode": AttendancePolicy.Mode.PART_TIME,
        "description": "按有效打卡天数核算，可通过人工调整补充小时。",
    },
    {
        "code": "shift",
        "name": "轮班",
        "mode": AttendancePolicy.Mode.SHIFT,
        "description": "按有效打卡天数核算，支持独立跨日截止时间。",
    },
]


class Command(BaseCommand):
    help = "初始化管理员、基础考勤策略和可选的参考表人员档案"

    def add_arguments(self, parser):
        parser.add_argument("--admin-username", default="admin")
        parser.add_argument("--admin-password", required=True)
        parser.add_argument("--reference", default="")

    def handle(self, *args, **options):
        username = options["admin_username"]
        password = options["admin_password"]
        if len(password) < 10:
            raise CommandError("管理员密码至少需要 10 位")
        user, created = User.objects.get_or_create(username=username, defaults={"is_staff": True, "is_superuser": True})
        user.is_staff = True
        user.is_superuser = True
        user.set_password(password)
        user.save()
        AccountProfile.objects.update_or_create(user=user, defaults={"role": AccountProfile.Role.ADMIN})
        self.stdout.write(self.style.SUCCESS(f"管理员账号 {username} 已{'创建' if created else '更新'}"))

        policy_map = {}
        for definition in POLICIES:
            policy, _ = AttendancePolicy.objects.update_or_create(
                code=definition["code"],
                defaults=definition,
            )
            policy_map[policy.code] = policy
        for name, color, description in [
            ("领导层", "#7C6FD1", "组织管理与决策岗位"),
            ("新员工", "#2B8CB8", "入职初期人员"),
            ("兼职", "#D58A25", "非全日制或临时人员"),
        ]:
            EmployeeTag.objects.update_or_create(name=name, defaults={"color": color, "description": description})
        self.stdout.write(self.style.SUCCESS("基础考勤策略和人员标签已就绪"))

        reference = options.get("reference")
        if reference:
            path = Path(reference)
            if not path.exists():
                raise CommandError(f"参考表不存在：{path}")
            count = self._import_reference(path, policy_map["standard"])
            self.stdout.write(self.style.SUCCESS(f"已从参考表导入/更新 {count} 条人员档案"))

    def _import_reference(self, path, default_policy):
        workbook = load_workbook(path, data_only=True, read_only=False)
        if "4月考勤汇总" not in workbook.sheetnames:
            raise CommandError("参考表中没有找到“4月考勤汇总”页")
        summary = workbook["4月考勤汇总"]
        raw_numbers = {}
        if "飞书打卡" in workbook.sheetnames:
            raw = workbook["飞书打卡"]
            for row in range(2, raw.max_row + 1):
                name = normalize_name(raw.cell(row, 1).value)
                if name:
                    raw_numbers.setdefault(name, cell_text(raw.cell(row, 3).value))

        leadership_tag = EmployeeTag.objects.get(name="领导层")
        part_time_tag = EmployeeTag.objects.get(name="兼职")
        current_department = ""
        count = 0
        for row in range(5, summary.max_row + 1):
            name = cell_text(summary.cell(row, 4).value)
            if not name:
                continue
            seq = int(summary.cell(row, 1).value or row - 4)
            department_cell = cell_text(summary.cell(row, 2).value)
            if department_cell:
                current_department = department_cell
            position = cell_text(summary.cell(row, 3).value)
            normalized_name = normalize_name(name)
            employee_no = raw_numbers.get(normalized_name) or f"HR{seq:04d}"
            status_text = cell_text(summary.cell(row, 6).value)
            status_map = {
                "试用期": Employee.EmploymentStatus.PROBATION,
                "已转正": Employee.EmploymentStatus.REGULAR,
                "创始人": Employee.EmploymentStatus.FOUNDER,
                "/": Employee.EmploymentStatus.FOUNDER,
            }
            join_value = summary.cell(row, 5).value
            join_date = self._to_date(join_value)
            employee, _ = Employee.objects.update_or_create(
                employee_no=employee_no,
                defaults={
                    "name": name.strip(),
                    "department": current_department,
                    "position": position,
                    "join_date": join_date,
                    "employment_status": status_map.get(status_text, Employee.EmploymentStatus.REGULAR),
                    "active": "离职" not in name,
                    "attendance_policy": default_policy,
                    "phone": cell_text(summary.cell(row, 16).value),
                    "bank_name": cell_text(summary.cell(row, 17).value),
                    "bank_account_holder": cell_text(summary.cell(row, 18).value),
                    "bank_province": cell_text(summary.cell(row, 19).value),
                    "bank_branch": cell_text(summary.cell(row, 20).value),
                    "bank_card_number": cell_text(summary.cell(row, 21).value),
                    "alipay_account": cell_text(summary.cell(row, 22).value),
                },
            )
            if any(keyword in position for keyword in ["董事长", "总经理", "负责人", "主管"]):
                employee.tags.add(leadership_tag)
            if "兼职" in position:
                employee.tags.add(part_time_tag)
            count += 1
        return count

    @staticmethod
    def _to_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return from_excel(value).date()
        return None

