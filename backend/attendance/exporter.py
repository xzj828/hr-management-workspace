import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import AttendancePolicy


THIN = Side(style="thin", color="9CA3AF")


def build_summary_workbook(batch):
    workbook = Workbook()
    summary = workbook.active
    summary.title = f"{batch.month}月考勤汇总"
    detail = workbook.create_sheet("核算明细")
    raw = workbook.create_sheet("原始打卡")

    days_in_month = calendar.monthrange(batch.year, batch.month)[1]
    summary.merge_cells("A1:W1")
    summary["A1"] = (
        f"{batch.month}月份共{days_in_month}天，默认应出勤{batch.default_expected_days}天。"
        "实际出勤由考勤策略、有效打卡和人工调整共同计算；请以核算明细为准。"
    )
    summary.merge_cells("A2:W2")
    summary["A2"] = f"西鸣科技-{batch.year}年{batch.month}月份考勤汇总表"
    summary["A2"].font = Font(size=16, bold=True)

    vertical_headers = {
        "A3:A4": "序号",
        "B3:B4": "部门",
        "C3:C4": "岗位",
        "D3:D4": "姓名",
        "E3:E4": "入职日期",
        "F3:F4": "状态",
        "G3:G4": "应出勤天数",
        "H3:H4": "休假天数",
        "I3:I4": "请假天数",
        "J3:J4": "加班时长\n（单位：天）",
        "K3:K4": "实际出勤天数",
        "P3:P4": "手机号码",
        "W3:W4": "核算备注",
    }
    for address, label in vertical_headers.items():
        summary.merge_cells(address)
        summary[address.split(":")[0]] = label
    summary.merge_cells("L3:O3")
    summary["L3"] = "考勤【主管核对】"
    summary.merge_cells("Q3:V3")
    summary["Q3"] = "银行卡信息"
    row4_headers = {
        "L4": "迟到/次",
        "M4": "旷工/次",
        "N4": "缺卡/次",
        "O4": "扣款",
        "Q4": "银行名称",
        "R4": "开户人",
        "S4": "开户省份",
        "T4": "开户行",
        "U4": "银行卡号",
        "V4": "支付宝账号",
    }
    for cell, value in row4_headers.items():
        summary[cell] = value

    detail_headers = [
        "工号",
        "姓名",
        "考勤模式",
        "应出勤",
        "有效打卡天数",
        "请假天数",
        "加班天数",
        "人工天数调整",
        "基础实际出勤",
        "实际出勤",
        "待审核跨日",
        "规则说明",
    ]
    detail.append(detail_headers)

    results = list(
        batch.results.select_related("employee", "employee__attendance_policy")
        .prefetch_related("employee__tags")
        .order_by("employee__department", "employee__employee_no")
    )
    for index, result in enumerate(results, start=1):
        employee = result.employee
        policy = employee.attendance_policy
        mode = policy.mode if policy else AttendancePolicy.Mode.STANDARD
        detail_row = index + 1
        detail.append(
            [
                employee.employee_no,
                employee.name,
                mode,
                float(result.due_days),
                float(result.punch_days),
                float(result.leave_days),
                float(result.overtime_days),
                float(result.adjustment_days),
                None,
                None,
                result.rule_trace.get("pending_cross_day_suspicions", 0),
                result.rule_trace.get("base_rule", ""),
            ]
        )
        detail[f"I{detail_row}"] = (
            f'=IF(OR(C{detail_row}="exempt",C{detail_row}="flexible"),D{detail_row},E{detail_row})'
        )
        detail[f"J{detail_row}"] = f"=ROUND(I{detail_row}+H{detail_row},2)"

        row = index + 4
        status_label = employee.get_employment_status_display()
        summary.append(
            [
                index,
                employee.department,
                employee.position,
                employee.name,
                employee.join_date,
                status_label,
                float(result.due_days),
                None,
                float(result.leave_days) or None,
                float(result.overtime_days) or None,
                None,
                result.late_count or None,
                result.absence_count or None,
                result.missing_punch_count or None,
                float(result.deduction) or None,
                employee.phone,
                employee.bank_name,
                employee.bank_account_holder,
                employee.bank_province,
                employee.bank_branch,
                employee.bank_card_number,
                employee.alipay_account,
                result.note,
            ]
        )
        summary[f"H{row}"] = f"={days_in_month}-G{row}"
        summary[f"K{row}"] = f"='核算明细'!J{detail_row}"

    raw_headers = ["姓名", "组织名称", "工号", "考勤规则"] + [str(day) for day in range(1, days_in_month + 1)]
    raw.append(raw_headers)
    grouped = {}
    for raw_day in batch.raw_days.all().order_by("source_row", "work_date"):
        row = grouped.setdefault(
            raw_day.source_row,
            [raw_day.source_name, raw_day.organization, raw_day.employee_no, raw_day.attendance_rule]
            + ["-"] * days_in_month,
        )
        row[3 + raw_day.work_date.day] = raw_day.raw_value or "-"
    for row in grouped.values():
        raw.append(row)

    _style_summary(summary, len(results) + 4)
    _style_table(detail, len(results) + 1, len(detail_headers))
    _style_table(raw, len(grouped) + 1, len(raw_headers))
    detail.sheet_state = "hidden"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _style_summary(sheet, max_row):
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    title_fill = PatternFill("solid", fgColor="F8FAFC")
    for cell in sheet[1]:
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=3, max_row=4, min_col=1, max_col=23):
        for cell in row:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="1E293B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in sheet.iter_rows(min_row=5, max_row=max_row, min_col=1, max_col=23):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column in {2, 3, 4, 20, 23} else "center")
    widths = [7, 15, 18, 11, 13, 11, 12, 11, 11, 13, 14, 9, 9, 9, 11, 15, 15, 12, 12, 30, 24, 20, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 32
    sheet.row_dimensions[3].height = 26
    sheet.row_dimensions[4].height = 32


def _style_table(sheet, max_row, max_col):
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="CBD5E1")
        cell.font = Font(bold=True, color="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 16 if column <= 12 else 10
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
