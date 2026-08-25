import tempfile
from datetime import date
from io import StringIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from recruitment.models import Candidate, RecruitmentJob

from .exporter import build_summary_workbook
from .models import AttendancePolicy, AttendanceResult, CrossDaySuspicion, Employee, ImportBatch, RawPunchDay
from .services import parse_punches, process_import_batch, resolve_cross_day


class LoginSessionTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="session-user", password="strong-password-123")

    def test_normal_login_expires_at_browser_close(self):
        response = self.client.post(
            "/api/auth/login/",
            {"username": self.user.username, "password": "strong-password-123", "remember": False},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.client.session.get_expire_at_browser_close())

    def test_remembered_login_uses_thirty_day_session(self):
        response = self.client.post(
            "/api/auth/login/",
            {"username": self.user.username, "password": "strong-password-123", "remember": True},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.client.session.get_expire_at_browser_close())
        self.assertGreaterEqual(self.client.session.get_expiry_age(), 29 * 24 * 60 * 60)


class AttendanceRuleTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser("admin", password="StrongPassword123!")
        self.standard = AttendancePolicy.objects.create(code="standard", name="标准考勤", mode="standard")
        self.exempt = AttendancePolicy.objects.create(code="exempt", name="免考勤", mode="exempt")
        self.employee = Employee.objects.create(
            employee_no="10001",
            name="测试员工",
            department="测试部",
            attendance_policy=self.standard,
        )
        self.exempt_employee = Employee.objects.create(
            employee_no="10002",
            name="免考勤员工",
            department="管理层",
            attendance_policy=self.exempt,
        )

    def make_upload(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "飞书打卡"
        sheet.append(["姓名", "组织名称", "工号", "考勤规则", "1", "2", "3", "4"])
        sheet.append(["测试员工", "测试部", "10001", "标准", "08:30\n23:40", "00:20", "-", "08:31\n18:00"])
        sheet.append(["免考勤员工", "管理层", "10002", "免考勤", "-", "-", "-", "-"])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            workbook.save(handle.name)
            data = Path(handle.name).read_bytes()
        return SimpleUploadedFile("attendance.xlsx", data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_parse_next_day_marker(self):
        punches = parse_punches("15:40\n次日00:13")
        self.assertEqual(len(punches), 2)
        self.assertEqual(punches[1]["minutes"], 1453)
        self.assertTrue(punches[1]["next_day"])

    def test_import_flags_single_early_punch_and_recalculates(self):
        upload = self.make_upload()
        batch = ImportBatch.objects.create(
            original_filename=upload.name,
            source_file=upload,
            file_sha256="a" * 64,
            year=2026,
            month=4,
            default_expected_days=3,
            uploaded_by=self.user,
        )
        process_import_batch(batch)
        suspicion = CrossDaySuspicion.objects.get(batch=batch)
        self.assertEqual(suspicion.work_date, date(2026, 4, 2))
        result = AttendanceResult.objects.get(batch=batch, employee=self.employee)
        self.assertEqual(float(result.punch_days), 2)
        self.assertEqual(result.status, AttendanceResult.Status.REVIEW)
        exempt_result = AttendanceResult.objects.get(batch=batch, employee=self.exempt_employee)
        self.assertEqual(float(exempt_result.actual_days), 3)
        resolve_cross_day(suspicion, CrossDaySuspicion.Status.KEEP_CURRENT, self.user)
        result.refresh_from_db()
        self.assertEqual(float(result.punch_days), 3)

    def test_export_contains_auditable_formula_and_raw_sheet(self):
        upload = self.make_upload()
        batch = ImportBatch.objects.create(
            original_filename=upload.name,
            source_file=upload,
            file_sha256="b" * 64,
            year=2026,
            month=4,
            default_expected_days=3,
            uploaded_by=self.user,
        )
        process_import_batch(batch)
        stream = build_summary_workbook(batch)
        workbook = load_workbook(stream, data_only=False)
        self.assertEqual(workbook.sheetnames, ["4月考勤汇总", "核算明细", "原始打卡"])
        self.assertTrue(str(workbook["4月考勤汇总"]["K5"].value).startswith("='核算明细'!"))
        self.assertTrue(str(workbook["核算明细"]["J2"].value).startswith("=ROUND"))

    def test_dashboard_aggregates_multiple_months(self):
        april = ImportBatch.objects.create(
            original_filename="april.xlsx",
            file_sha256="c" * 64,
            year=2026,
            month=4,
            default_expected_days=25,
            status=ImportBatch.Status.COMPLETED,
            total_rows=2,
            matched_rows=2,
            uploaded_by=self.user,
        )
        may = ImportBatch.objects.create(
            original_filename="may.xlsx",
            file_sha256="d" * 64,
            year=2026,
            month=5,
            default_expected_days=25,
            status=ImportBatch.Status.COMPLETED,
            total_rows=2,
            matched_rows=2,
            uploaded_by=self.user,
        )
        for batch, standard_actual in ((april, 20), (may, 22)):
            AttendanceResult.objects.create(
                batch=batch,
                employee=self.employee,
                due_days=25,
                actual_days=standard_actual,
                status=AttendanceResult.Status.REVIEW,
            )
            AttendanceResult.objects.create(
                batch=batch,
                employee=self.exempt_employee,
                due_days=25,
                actual_days=25,
                status=AttendanceResult.Status.NORMAL,
            )
            RawPunchDay.objects.create(
                batch=batch,
                employee=self.employee,
                employee_no=self.employee.employee_no,
                source_name=self.employee.name,
                source_row=2,
                work_date=date(batch.year, batch.month, 1),
                has_punch=True,
                effective_has_punch=True,
            )

        self.client.force_login(self.user)
        response = self.client.get("/api/dashboard/?from=2026-04&to=2026-05")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["period"]["batch_count"], 2)
        self.assertEqual(payload["kpis"]["employees"], 2)
        self.assertEqual(payload["kpis"]["attendance_rate"], 92.0)
        self.assertEqual(payload["kpis"]["review_count"], 1)
        self.assertEqual(len(payload["daily"]), 2)
        self.assertEqual(response.json()["summary"]["total_rows"], 4)

        filtered = self.client.get("/api/dashboard/?from=2026-04&to=2026-05&department=测试部")
        self.assertEqual(filtered.status_code, 200)
        filtered_payload = filtered.json()
        self.assertEqual(filtered_payload["selected_department"], "测试部")
        self.assertEqual(filtered_payload["kpis"]["employees"], 1)
        self.assertEqual(filtered_payload["kpis"]["attendance_rate"], 84.0)
        self.assertEqual(filtered_payload["available_departments"], ["测试部", "管理层"])

        invalid = self.client.get("/api/dashboard/?from=2026-06&to=2026-05")
        self.assertEqual(invalid.status_code, 400)


class DemoWorkspaceCommandTests(TestCase):
    def setUp(self):
        self.temp_media = tempfile.TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.temp_media.name)
        self.media_override.enable()
        call_command(
            "setup_system",
            admin_username="admin",
            admin_password="Demo-command-2026!",
            stdout=StringIO(),
        )

    def tearDown(self):
        self.media_override.disable()
        self.temp_media.cleanup()

    def test_command_is_idempotent_and_clearable(self):
        call_command("load_demo_workspace", stdout=StringIO())
        first_counts = {
            "employees": Employee.objects.filter(employee_no__startswith="DEMO-").count(),
            "batches": ImportBatch.objects.filter(
                original_filename__startswith="演示考勤-"
            ).count(),
            "results": AttendanceResult.objects.filter(
                batch__original_filename__startswith="演示考勤-"
            ).count(),
            "suspicions": CrossDaySuspicion.objects.filter(
                batch__original_filename__startswith="演示考勤-"
            ).count(),
            "jobs": RecruitmentJob.objects.filter(is_demo=True).count(),
            "candidates": Candidate.objects.filter(is_demo=True).count(),
        }
        self.assertEqual(first_counts["employees"], 8)
        self.assertGreaterEqual(first_counts["batches"], 1)
        self.assertEqual(first_counts["results"], first_counts["batches"] * 8)
        self.assertEqual(first_counts["suspicions"], first_counts["batches"])
        self.assertEqual(first_counts["jobs"], 3)
        self.assertEqual(first_counts["candidates"], 10)

        call_command("load_demo_workspace", stdout=StringIO())
        self.assertEqual(
            ImportBatch.objects.filter(
                original_filename__startswith="演示考勤-"
            ).count(),
            first_counts["batches"],
        )
        self.assertEqual(Employee.objects.filter(employee_no__startswith="DEMO-").count(), 8)
        self.assertEqual(RecruitmentJob.objects.filter(is_demo=True).count(), 3)
        self.assertEqual(Candidate.objects.filter(is_demo=True).count(), 10)

        call_command("load_demo_workspace", clear=True, stdout=StringIO())
        self.assertFalse(Employee.objects.filter(employee_no__startswith="DEMO-").exists())
        self.assertFalse(
            ImportBatch.objects.filter(
                original_filename__startswith="演示考勤-"
            ).exists()
        )
        self.assertFalse(RecruitmentJob.objects.filter(is_demo=True).exists())
        self.assertFalse(Candidate.objects.filter(is_demo=True).exists())

    def test_command_refuses_to_mix_with_real_attendance_data(self):
        policy = AttendancePolicy.objects.get(code="standard")
        Employee.objects.create(
            employee_no="REAL-001",
            name="真实员工",
            department="测试部",
            join_date=date(2026, 1, 1),
            attendance_policy=policy,
        )

        with self.assertRaisesMessage(CommandError, "拒绝混合加载"):
            call_command("load_demo_workspace", stdout=StringIO())

        self.assertFalse(Employee.objects.filter(employee_no__startswith="DEMO-").exists())
        self.assertFalse(ImportBatch.objects.exists())

    def test_failed_reload_preserves_previous_attendance_batches(self):
        call_command("load_demo_workspace", stdout=StringIO())
        previous_ids = set(
            ImportBatch.objects.filter(
                original_filename__startswith="演示考勤-"
            ).values_list("id", flat=True)
        )
        previous_files = {
            path.relative_to(self.temp_media.name).as_posix()
            for path in Path(self.temp_media.name).rglob("*.xlsx")
        }

        with patch(
            "attendance.management.commands.load_demo_workspace.load_demo_data",
            side_effect=RuntimeError("simulated recruitment failure"),
        ):
            with self.assertRaisesMessage(RuntimeError, "simulated recruitment failure"):
                call_command("load_demo_workspace", stdout=StringIO())

        self.assertEqual(
            set(
                ImportBatch.objects.filter(
                    original_filename__startswith="演示考勤-"
                ).values_list("id", flat=True)
            ),
            previous_ids,
        )
        self.assertEqual(
            {
                path.relative_to(self.temp_media.name).as_posix()
                for path in Path(self.temp_media.name).rglob("*.xlsx")
            },
            previous_files,
        )

    def test_month_start_still_creates_two_periods_and_one_pending_review(self):
        with patch(
            "attendance.management.commands.load_demo_workspace.timezone.localdate",
            return_value=date(2026, 8, 3),
        ):
            call_command("load_demo_workspace", stdout=StringIO())

        self.assertEqual(
            set(
                ImportBatch.objects.filter(
                    original_filename__startswith="演示考勤-"
                ).values_list("year", "month")
            ),
            {(2026, 6), (2026, 7)},
        )
        self.assertEqual(
            CrossDaySuspicion.objects.filter(
                status=CrossDaySuspicion.Status.PENDING
            ).count(),
            1,
        )
